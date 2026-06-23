from __future__ import annotations

import argparse
import re
from collections import OrderedDict
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_TEMPLATE = "ficheiro template 2.xlsx"
DEFAULT_OUTPUT = "simulacao-final.xlsx"

COMPETITOR_COLUMNS = ("CONTINENTE", "LIDL", "PINGO-DOCE")
DADOS_HEADER_ROW = 2
EURO_FORMAT = '#,##0.00 "€"'

PROPOSTA_OPTIONS = ("Subir", "Descer", "OK")
FEEDBACK_OPTIONS = (
    "OK Envio ao ficheiro",
    "NOK",
    "Sem Con",
    "Nego",
    "Acompanhar",
)

PROPOSTA_LIST_COL = 4  # Folha2!D
FEEDBACK_LIST_COL = 6  # Folha2!F


def build_header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {as_text(value).upper(): index for index, value in enumerate(row) if as_text(value)}


def required_column(headers: dict[str, int], name: str) -> int:
    key = name.strip().upper()
    if key not in headers:
        raise ValueError(f"Header not found: {name}")
    return headers[key]


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def item_key(value: Any) -> str:
    text = as_text(value)
    return text.lstrip("0") or text


def ean_key(value: Any) -> str:
    text = as_text(value)
    return text.lstrip("0") or text


def excluded_article(value: Any) -> bool:
    description = as_text(value).upper()
    return description.startswith(("SUB", "XXX"))


def parse_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return datetime(value.year, value.month, value.day)
    text = as_text(value)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def format_slash_date(value: Any) -> str:
    parsed = parse_date(value)
    if parsed is None:
        return as_text(value)
    return parsed.strftime("%d/%m/%Y")


def extract_week_number(filename: str) -> int | None:
    match = re.search(r"(?i)(?:^|[^A-Z0-9])S(\d{1,2})(?:[^0-9]|$)", filename)
    return int(match.group(1)) if match else None


def extract_date_yyyymmdd(filename: str) -> datetime | None:
    match = re.search(r"(\d{8})", filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y%m%d")
    except ValueError:
        return None


def extract_date_dd_mm_yyyy(filename: str) -> datetime | None:
    match = re.search(r"(\d{2}-\d{2}-\d{4})", filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%d-%m-%Y")
    except ValueError:
        return None


def find_latest_simulation(path: Path) -> Path:
    candidates: list[tuple[int, Path]] = []
    for candidate in path.glob("*Simula*.xls*"):
        week = extract_week_number(candidate.name)
        if week is not None:
            candidates.append((week, candidate))
    if not candidates:
        raise FileNotFoundError("Nenhum ficheiro de Simulação PVP encontrado.")
    return max(candidates, key=lambda item: item[0])[1]


def find_analise_precos(path: Path, week: int) -> Path | None:
    candidates: list[Path] = []
    for candidate in path.glob("*.xls*"):
        if extract_week_number(candidate.name) != week:
            continue
        name_lower = candidate.name.lower()
        if "suivi" in name_lower:
            continue
        if "analise" in name_lower or "análise" in name_lower or "precos" in name_lower or "preços" in name_lower:
            candidates.append(candidate)
    return sorted(candidates)[-1] if candidates else None


def find_latest_comparavel(path: Path) -> Path:
    candidates: list[tuple[datetime, Path]] = []
    for candidate in path.glob("*compar*"):
        date = extract_date_yyyymmdd(candidate.name)
        if date is not None:
            candidates.append((date, candidate))
    if not candidates:
        raise FileNotFoundError("Nenhum Relatório comparável encontrado.")
    return max(candidates, key=lambda item: item[0])[1]


def find_latest_total_meas(path: Path) -> Path:
    candidates: list[tuple[datetime, Path]] = []
    for candidate in path.glob("TOTAL*meas*"):
        date = extract_date_dd_mm_yyyy(candidate.name)
        if date is not None:
            candidates.append((date, candidate))
    if not candidates:
        raise FileNotFoundError("Nenhum ficheiro TOTAL - meas encontrado.")
    return max(candidates, key=lambda item: item[0])[1]


def output_filename_for_week(week: int) -> str:
    return f"S{week} - Análise preços SUIVI.xlsx"


def load_simulation(simulation_path: Path) -> tuple[list[dict[str, Any]], datetime | None]:
    wb = load_workbook(simulation_path, data_only=True, read_only=True, keep_vba=True)
    dados_ws = wb["Dados"]
    shopping_date = parse_date(dados_ws["G1"].value) or parse_date(dados_ws["H1"].value)
    headers = build_header_map(
        next(dados_ws.iter_rows(min_row=DADOS_HEADER_ROW, max_row=DADOS_HEADER_ROW, values_only=True))
    )
    cols = {
        "itm8": required_column(headers, "ITM8"),
        "ean": required_column(headers, "EAN"),
        "description": required_column(headers, "Descrição"),
        "brand": required_column(headers, "Marca"),
        "pvp_competitor": required_column(headers, "PVP Concorrente"),
        "pvp_future": required_column(headers, "PVP Cadencier Futuro"),
        "competitor": required_column(headers, "Insignia Concorrente"),
        "psycho": required_column(headers, "Psyco"),
        "tipo": required_column(headers, "Tipo Produto"),
        "argus": required_column(headers, "Argus"),
        "aval": required_column(headers, "Aval"),
        "amont": required_column(headers, "Amont"),
        "st": required_column(headers, "Estatuto"),
        "promo_perm": required_column(headers, "Promo Permanente"),
        "edlp": required_column(headers, "EDLP"),
    }

    records: OrderedDict[tuple[Any, ...], dict[str, Any]] = OrderedDict()
    price_totals: dict[tuple[Any, ...], dict[str, list[float]]] = {}
    for row in dados_ws.iter_rows(min_row=DADOS_HEADER_ROW + 1, values_only=True):
        itm8 = as_text(row[cols["itm8"]])
        ean = as_text(row[cols["ean"]])
        if not itm8 and not ean:
            continue
        if excluded_article(row[cols["description"]]):
            continue

        competitor = as_text(row[cols["competitor"]]).upper()
        if competitor not in COMPETITOR_COLUMNS:
            continue

        key = (
            row[cols["amont"]],
            row[cols["aval"]],
            itm8,
            ean,
            row[cols["description"]],
            row[cols["brand"]],
            row[cols["tipo"]],
            row[cols["psycho"]],
            row[cols["argus"]],
            row[cols["promo_perm"]],
            row[cols["edlp"]],
            row[cols["pvp_future"]],
        )
        if key not in records:
            records[key] = {
                "amont": row[cols["amont"]],
                "aval": row[cols["aval"]],
                "itm8": itm8,
                "description": row[cols["description"]],
                "brand": row[cols["brand"]],
                "ean": ean,
                "tipo": row[cols["tipo"]],
                "psycho": row[cols["psycho"]],
                "promo_perm": row[cols["promo_perm"]],
                "edlp": row[cols["edlp"]],
                "argus": row[cols["argus"]],
                "st": row[cols["st"]],
                "pvp_cadencier": row[cols["pvp_future"]],
                "prices": {},
            }
            price_totals[key] = {}

        price = as_number(row[cols["pvp_competitor"]])
        if price is not None:
            price_totals[key].setdefault(competitor, []).append(price)

    for key, totals in price_totals.items():
        records[key]["prices"] = {
            competitor: sum(values) / len(values)
            for competitor, values in totals.items()
            if values
        }

    wb.close()

    sorted_records = sorted(
        records.values(),
        key=lambda record: (
            as_text(record["amont"])[1:3],
            as_text(record["amont"])[4:7],
            as_text(record["aval"])[4:6],
            as_text(record["aval"])[7:9],
            as_text(record["brand"]),
            as_number(record["pvp_cadencier"]) if as_number(record["pvp_cadencier"]) is not None else 999999999,
        ),
    )
    return sorted_records, shopping_date


def load_comparavel(comparavel_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(comparavel_path, data_only=True, read_only=True)
    ws = wb["Produtos"]

    promos: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        ean = as_text(row[0])
        if not ean:
            continue
        promos[ean] = {
            "CONTINENTE": row[17] if row[17] is not None else "",
            "LIDL": row[25] if row[25] is not None else "",
            "PINGO-DOCE": row[21] if row[21] is not None else "",
        }
        normalized = ean_key(ean)
        if normalized and normalized not in promos:
            promos[normalized] = promos[ean]

    wb.close()
    return promos


def header_index(headers: tuple[Any, ...], name: str) -> int:
    normalized_name = name.strip().upper()
    for index, value in enumerate(headers):
        if as_text(value).upper() == normalized_name:
            return index
    raise ValueError(f"Header not found: {name}")


def find_header_column(headers: tuple[Any, ...], *keywords: str, exclude: tuple[str, ...] = ()) -> int | None:
    for index, value in enumerate(headers):
        label = as_text(value).upper()
        if any(excluded in label for excluded in exclude):
            continue
        if all(keyword in label for keyword in keywords):
            return index
    return None


def read_total_meas_rows(total_meas_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(total_meas_path, data_only=True, read_only=True)
    ws = wb["sql_query3"]
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    cols = {
        "grupo": header_index(headers, "GRUPO_INTERNO"),
        "uvc": header_index(headers, "UVC"),
        "ean": header_index(headers, "EAN"),
        "descricao": header_index(headers, "DESCRIÇÃO"),
        "in_mea": header_index(headers, "IN_MEA"),
        "pvp": header_index(headers, "PVP"),
    }

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date = parse_date(row[cols["in_mea"]])
        if date is None:
            continue
        rows.append(
            {
                "grupo": row[cols["grupo"]],
                "uvc": as_text(row[cols["uvc"]]),
                "ean": as_text(row[cols["ean"]]),
                "descricao": row[cols["descricao"]],
                "in_mea": date,
                "pvp": row[cols["pvp"]] if row[cols["pvp"]] is not None else "",
            }
        )
    wb.close()
    return rows


def sort_meas_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            as_text(row["grupo"]),
            as_text(row["uvc"]).zfill(10),
            ean_key(row["ean"]),
            as_text(row["descricao"]),
            -row["in_mea"].timestamp(),
        ),
    )


def build_meas_processed_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sorted_rows = sort_meas_rows(rows)
    best_by_ean: dict[str, dict[str, Any]] = {}
    for row in sorted_rows:
        key = ean_key(row["ean"]) or as_text(row["ean"])
        if not key:
            continue
        current = best_by_ean.get(key)
        if current is None or row["in_mea"] > current["in_mea"]:
            best_by_ean[key] = row

    return sort_meas_rows(list(best_by_ean.values()))


def build_meas_lookup(processed_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for row in processed_rows:
        entry = {
            "sort_date": row["in_mea"],
            "date": row["in_mea"],
            "pvp": row["pvp"],
            "uvc": row["uvc"],
            "ean": row["ean"],
        }
        for key in (as_text(row["ean"]), ean_key(row["ean"])):
            if key:
                lookup[key] = entry
    return lookup


def meas_lookup_for_record(lookup: dict[str, dict[str, Any]], record: dict[str, Any]) -> dict[str, Any]:
    for key in (as_text(record["ean"]), ean_key(record["ean"])):
        if key in lookup:
            return lookup[key]
    return {}


def write_total_meas_processed(total_meas_path: Path, pivot_rows: list[dict[str, Any]], processed_rows: list[dict[str, Any]]) -> Path:
    output_path = total_meas_path.with_name(f"{total_meas_path.stem} - processado.xlsx")
    wb = load_workbook(total_meas_path)
    if "TD Meas" in wb.sheetnames:
        del wb["TD Meas"]
    if "Meas Processado" in wb.sheetnames:
        del wb["Meas Processado"]

    pivot_ws = wb.create_sheet("TD Meas")
    processed_ws = wb.create_sheet("Meas Processado")
    headers = ("GRUPO_INTERNO", "UVC", "EAN", "DESCRIÇÃO", "IN_MEA", "PVP")

    for ws in (pivot_ws, processed_ws):
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col).value = header

    for target_ws, source_rows in ((pivot_ws, pivot_rows), (processed_ws, processed_rows)):
        for row_number, row in enumerate(source_rows, start=2):
            target_ws.cell(row_number, 1).value = row["grupo"]
            target_ws.cell(row_number, 2).value = row["uvc"]
            target_ws.cell(row_number, 3).value = row["ean"]
            target_ws.cell(row_number, 4).value = row["descricao"]
            date_cell = target_ws.cell(row_number, 5)
            date_cell.value = row["in_mea"]
            date_cell.number_format = "dd-mm-yyyy"
            target_ws.cell(row_number, 6).value = row["pvp"]

    wb.save(output_path)
    wb.close()
    return output_path


def count_meas_matches(records: list[dict[str, Any]], lookup: dict[str, dict[str, Any]]) -> int:
    return sum(1 for record in records if meas_lookup_for_record(lookup, record))


def load_total_meas(
    total_meas_path: Path,
    *,
    save_processed: bool = True,
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    raw_rows = read_total_meas_rows(total_meas_path)
    pivot_rows = sort_meas_rows(raw_rows)
    processed_rows = build_meas_processed_rows(raw_rows)
    lookup = build_meas_lookup(processed_rows)

    stats = {
        "raw_rows": len(raw_rows),
        "pivot_rows": len(pivot_rows),
        "processed_ean": len(processed_rows),
    }

    if save_processed:
        stats["processed_path"] = write_total_meas_processed(total_meas_path, pivot_rows, processed_rows)

    return lookup, stats


def load_analise_historico(analise_path: Path | None) -> dict[str, dict[str, dict[str, Any]]]:
    empty: dict[str, dict[str, Any]] = {"itm8": {}, "ean": {}}
    if analise_path is None or not analise_path.exists():
        return {"suivi": dict(empty), "comercial": dict(empty)}

    wb = load_workbook(analise_path, data_only=True, read_only=True)
    ws = wb["Folha1"]

    header_row = None
    headers: tuple[Any, ...] | None = None
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        labels = [as_text(value).upper() for value in row]
        if "ITM8" in labels:
            header_row = row_number
            headers = row
            break

    if header_row is None or headers is None:
        wb.close()
        raise ValueError(f"Não foi possível encontrar cabeçalhos em {analise_path}")

    itm8_col = next(index for index, value in enumerate(headers) if as_text(value).upper() == "ITM8")
    ean_col = next((index for index, value in enumerate(headers) if as_text(value).upper() == "EAN"), None)
    suivi_col = find_header_column(headers, "COMENT", "SUIVI", exclude=("HIST",))
    comercial_col = find_header_column(headers, "COMENT", "COMERCIAL", exclude=("HIST", "FEEDBACK"))

    result = {"suivi": dict(empty), "comercial": dict(empty)}
    column_map = {
        "suivi": suivi_col,
        "comercial": comercial_col,
    }

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        itm8 = item_key(row[itm8_col])
        ean = ean_key(row[ean_col]) if ean_col is not None else ""
        for key, col_index in column_map.items():
            if col_index is None:
                continue
            value = row[col_index] if row[col_index] is not None else ""
            if itm8 and itm8 not in result[key]["itm8"]:
                result[key]["itm8"][itm8] = value
            if ean and ean not in result[key]["ean"]:
                result[key]["ean"][ean] = value

    wb.close()
    return result


def historic_value(historic: dict[str, dict[str, Any]], record: dict[str, Any]) -> Any:
    by_itm8 = historic.get("itm8", {})
    by_ean = historic.get("ean", {})
    return by_itm8.get(item_key(record["itm8"]), by_ean.get(ean_key(record["ean"]), ""))


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def prepare_template(template_path: Path):
    wb = load_workbook(template_path)
    ws = wb["Folha1"]
    setup_folha2_lists(wb["Folha2"])
    setup_data_validations(ws)
    return wb, ws


def setup_folha2_lists(ws2) -> None:
    ws2.cell(1, PROPOSTA_LIST_COL).value = "PROPOSTA"
    for index, option in enumerate(PROPOSTA_OPTIONS, start=2):
        ws2.cell(index, PROPOSTA_LIST_COL).value = option

    ws2.cell(1, FEEDBACK_LIST_COL).value = "FEEDBACK COMERCIAL"
    for index, option in enumerate(FEEDBACK_OPTIONS, start=2):
        ws2.cell(index, FEEDBACK_LIST_COL).value = option


def setup_data_validations(ws) -> None:
    ws.data_validations.dataValidation.clear()
    last_row = max(ws.max_row, 5000)

    proposta_col = get_column_letter(33)  # AG
    feedback_col = get_column_letter(36)  # AJ
    proposta_range = f"Folha2!${get_column_letter(PROPOSTA_LIST_COL)}$2:${get_column_letter(PROPOSTA_LIST_COL)}${1 + len(PROPOSTA_OPTIONS)}"
    feedback_range = f"Folha2!${get_column_letter(FEEDBACK_LIST_COL)}$2:${get_column_letter(FEEDBACK_LIST_COL)}${1 + len(FEEDBACK_OPTIONS)}"

    proposta_validation = DataValidation(type="list", formula1=proposta_range, allow_blank=True)
    feedback_validation = DataValidation(type="list", formula1=feedback_range, allow_blank=True)
    proposta_validation.add(f"{proposta_col}4:{proposta_col}{last_row}")
    feedback_validation.add(f"{feedback_col}4:{feedback_col}{last_row}")
    ws.add_data_validation(proposta_validation)
    ws.add_data_validation(feedback_validation)


def clear_data_area(ws, max_rows: int) -> None:
    last_row = max(ws.max_row, max_rows)
    for row in ws.iter_rows(min_row=4, max_row=last_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def as_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = as_text(value).replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def promo_lookup(promos: dict[str, dict[str, Any]], ean: str) -> dict[str, Any]:
    if ean in promos:
        return promos[ean]
    normalized = ean_key(ean)
    return promos.get(normalized, {})


def apply_row_formulas(ws, row_number: int) -> None:
    row = row_number
    ws.cell(row, 3).value = f"=MID(A{row},2,2)"
    ws.cell(row, 4).value = f"=XLOOKUP(C{row},Folha2!A:A,Folha2!B:B)"
    ws.cell(row, 5).value = f"=MID(A{row},5,3)"
    ws.cell(row, 6).value = f"=MID(B{row},5,2)"
    ws.cell(row, 7).value = f"=MID(B{row},8,2)"
    ws.cell(row, 25).value = (
        f'=IF(OR(N{row}="O",O{row}="O"),IF(COUNT(V{row}:X{row})>0,MIN(V{row}:X{row}),MIN(S{row}:U{row})),'
        f'IF(M{row}="O",MIN(S{row}:U{row}),IF(L{row}="E",IFERROR(MODE(S{row}:U{row}),MIN(S{row}:U{row})),AVERAGE(S{row}:U{row}))))'
    )
    ws.cell(row, 26).value = f"=R{row}/Y{row}-1"
    ws.cell(row, 27).value = f"=R{row}-Y{row}"
    ws.cell(row, 28).value = f'=IF(R{row}=Y{row},"VERDADEIRO","FALSO")'
    ws.cell(row, 31).value = f'=IF(AC{row}="","",DATE(YEAR(AC{row}),MONTH(AC{row}),DAY(AC{row})))'


def fill_row(
    ws,
    row_number: int,
    record: dict[str, Any],
    promos: dict[str, Any],
    meas: dict[str, Any],
    historico_suivi: Any,
    historico_comercial: Any,
) -> None:
    if not as_text(record["ean"]):
        return

    base_values = {
        1: record["amont"],
        2: record["aval"],
        8: record["itm8"],
        9: record["description"],
        10: record["brand"],
        11: record["ean"],
        12: record["tipo"],
        13: record["psycho"],
        14: record["promo_perm"],
        15: record["edlp"],
        16: record["argus"],
        17: record["st"],
        18: record["pvp_cadencier"],
    }
    for col, value in base_values.items():
        ws.cell(row_number, col).value = value

    prices = record["prices"]
    ws.cell(row_number, 19).value = prices.get("CONTINENTE")
    ws.cell(row_number, 20).value = prices.get("LIDL")
    ws.cell(row_number, 21).value = prices.get("PINGO-DOCE")

    promo_values = promo_lookup(promos, as_text(record["ean"]))
    ws.cell(row_number, 22).value = promo_values.get("CONTINENTE", "")
    ws.cell(row_number, 23).value = promo_values.get("LIDL", "")
    ws.cell(row_number, 24).value = promo_values.get("PINGO-DOCE", "")

    meas_values = meas_lookup_for_record(meas, record)
    campaign_date_cell = ws.cell(row_number, 29)
    campaign_date_cell.value = meas_values.get("date", "")
    if campaign_date_cell.value not in (None, ""):
        campaign_date_cell.number_format = "dd-mm-yyyy"
    ws.cell(row_number, 30).value = meas_values.get("pvp", "")

    for col in range(19, 26):
        ws.cell(row_number, col).number_format = EURO_FORMAT

    apply_row_formulas(ws, row_number)

    ws.cell(row_number, 35).value = historico_suivi if historico_suivi is not None else ""
    ws.cell(row_number, 38).value = historico_comercial if historico_comercial is not None else ""


def apply_row_styles(ws, rows_count: int) -> None:
    max_col = ws.max_column
    source_cells = [ws.cell(4, col) for col in range(1, max_col + 1)]
    for row_number in range(4, rows_count + 4):
        for col in range(1, max_col + 1):
            copy_cell_style(source_cells[col - 1], ws.cell(row_number, col))


def build_workbook(
    template_path: Path,
    simulation_path: Path,
    comparavel_path: Path,
    total_meas_path: Path,
    analise_path: Path | None,
    output_path: Path,
) -> dict[str, Any]:
    records, shopping_date = load_simulation(simulation_path)
    comparavel = load_comparavel(comparavel_path)
    total_meas, meas_stats = load_total_meas(total_meas_path)
    historico = load_analise_historico(analise_path)

    wb, ws = prepare_template(template_path)
    clear_data_area(ws, len(records) + 3)
    apply_row_styles(ws, len(records))

    date_label = format_slash_date(shopping_date)
    if date_label:
        ws["S2"] = f"Shopping {date_label}"
        ws["V2"] = f"PROMO {date_label}"

    for offset, record in enumerate(records, start=4):
        fill_row(
            ws,
            offset,
            record,
            comparavel,
            total_meas,
            historic_value(historico["suivi"], record),
            historic_value(historico["comercial"], record),
        )

    first_extra_row = len(records) + 4
    if ws.max_row >= first_extra_row:
        ws.delete_rows(first_extra_row, ws.max_row - first_extra_row + 1)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output_path)

    meas_stats["matched_rows"] = count_meas_matches(records, total_meas)
    meas_stats["total_rows"] = len(records)
    return meas_stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera o ficheiro Análise preços SUIVI a partir do template e ficheiros de origem.")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE)
    parser.add_argument("--simulation", default=None)
    parser.add_argument("--comparavel", default=None)
    parser.add_argument("--total-meas", default=None)
    parser.add_argument("--analise-precos", default=None, help="Ficheiro SXX - analise preços da semana anterior.")
    parser.add_argument("--output", default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = Path.cwd()

    simulation_path = Path(args.simulation) if args.simulation else find_latest_simulation(root)
    week = extract_week_number(simulation_path.name)
    if week is None:
        raise ValueError(f"Não foi possível extrair a semana de {simulation_path.name}")

    comparavel_path = Path(args.comparavel) if args.comparavel else find_latest_comparavel(root)
    total_meas_path = Path(args.total_meas) if args.total_meas else find_latest_total_meas(root)
    analise_path = (
        Path(args.analise_precos)
        if args.analise_precos
        else find_analise_precos(root, week - 1)
    )
    output_path = Path(args.output) if args.output else root / output_filename_for_week(week)

    meas_stats = build_workbook(
        root / args.template,
        simulation_path,
        comparavel_path,
        total_meas_path,
        analise_path,
        output_path,
    )

    print(f"Ficheiro criado: {output_path.name}")
    print(f"Simulação: {simulation_path.name}")
    print(f"Comparável: {comparavel_path.name}")
    print(f"Total meas: {total_meas_path.name}")
    if "processed_path" in meas_stats:
        print(f"Total meas processado: {meas_stats['processed_path'].name}")
    print(
        "Campanha/PVP preenchidos (EAN): "
        f"{meas_stats['matched_rows']}/{meas_stats['total_rows']} "
        f"({100 * meas_stats['matched_rows'] / meas_stats['total_rows']:.1f}%)"
    )
    if analise_path is None:
        print("Aviso: ficheiro de análise preços da semana anterior não encontrado; históricos ficaram vazios.")
    else:
        print(f"Análise preços (S{week - 1}): {analise_path.name}")


if __name__ == "__main__":
    main()
